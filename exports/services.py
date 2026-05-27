from dataclasses import dataclass
from datetime import date as date_class
from io import BytesIO

from django.core.exceptions import PermissionDenied, ValidationError
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from accounts.models import UserRole
from control.services import build_plan_rows, group_rows_by_organization, plans_for_user, split_summary
from questionnaires.models import (
    CheckStatus,
    MessengerCheckStatus,
    MessengerFormStatus,
    MessengerTestForm,
    MessengerTestResult,
    VpnFormStatus,
    VpnTestForm,
    VpnMeasurementResult,
)


FORM_TYPE_ALL = "all"
FORM_TYPE_VPN = "vpn"
FORM_TYPE_MESSENGER = "messenger"

RESULT_LABELS = {
    CheckStatus.WORKS: "Работает",
    CheckStatus.NOT_WORKS: "Не работает",
    CheckStatus.UNABLE_TO_CHECK: "Невозможно проверить",
    CheckStatus.NOT_APPLICABLE: "-",
    MessengerCheckStatus.WORKS: "Работает",
    MessengerCheckStatus.NOT_WORKS: "Не работает",
    MessengerCheckStatus.UNABLE_TO_CHECK: "Невозможно проверить",
    "": "",
    None: "",
}

FORM_STATUS_LABELS = {
    VpnFormStatus.DRAFT: "Черновик",
    VpnFormStatus.PARTIALLY_SUBMITTED: "Частично отправлена",
    VpnFormStatus.SUBMITTED: "Отправлена",
    VpnFormStatus.CONFIRMED: "Подтверждена",
    VpnFormStatus.RETURNED: "Возвращена",
    MessengerFormStatus.DRAFT: "Черновик",
    MessengerFormStatus.SUBMITTED: "Отправлена",
    MessengerFormStatus.CONFIRMED: "Подтверждена",
    MessengerFormStatus.RETURNED: "Возвращена",
}


@dataclass
class ExportFilters:
    date_from: date_class
    date_to: date_class
    form_type: str = FORM_TYPE_ALL
    organization_id: int | None = None
    city_id: int | None = None
    mobile_operator_id: int | None = None
    os: str = ""
    status: str = ""
    vpn_form_ids: tuple[int, ...] = ()
    messenger_form_ids: tuple[int, ...] = ()
    selection_mode: bool = False

    @property
    def is_single_date(self):
        return self.date_from == self.date_to

    def as_dict(self):
        return {
            "date_from": self.date_from.isoformat(),
            "date_to": self.date_to.isoformat(),
            "form_type": self.form_type,
            "organization": self.organization_id,
            "city": self.city_id,
            "mobile_operator": self.mobile_operator_id,
            "os": self.os,
            "status": self.status,
            "vpn_forms": list(self.vpn_form_ids),
            "messenger_forms": list(self.messenger_form_ids),
            "selection_mode": self.selection_mode,
        }


@dataclass
class ExcelExportResult:
    content: bytes
    file_name: str
    rows_vpn: int
    rows_messenger: int
    ready_for_export: bool


def can_export(user) -> bool:
    return user.is_authenticated and (
        user.is_superuser or user.role in [UserRole.ADMIN, UserRole.READER, UserRole.COORDINATOR]
    )


def parse_export_filters(params, user) -> ExportFilters:
    if not can_export(user):
        raise PermissionDenied

    date_value = params.get("date")
    date_from_value = params.get("date_from")
    date_to_value = params.get("date_to")

    if date_value:
        date_from = date_to = _parse_date(date_value)
    elif date_from_value or date_to_value:
        date_from = _parse_date(date_from_value)
        date_to = _parse_date(date_to_value or date_from_value)
    else:
        date_from = date_to = timezone.localdate()

    if date_to < date_from:
        raise ValidationError("Дата окончания не может быть меньше даты начала.")

    form_type = params.get("form_type") or FORM_TYPE_ALL
    if form_type not in [FORM_TYPE_ALL, FORM_TYPE_VPN, FORM_TYPE_MESSENGER]:
        form_type = FORM_TYPE_ALL

    organization_id = _optional_int(params.get("organization"))
    if user.role == UserRole.COORDINATOR and not user.is_superuser:
        organization_id = user.organization_id

    return ExportFilters(
        date_from=date_from,
        date_to=date_to,
        form_type=form_type,
        organization_id=organization_id,
        city_id=_optional_int(params.get("city")),
        mobile_operator_id=_optional_int(params.get("mobile_operator")),
        os=params.get("os") or "",
        status=params.get("status") or "",
        vpn_form_ids=tuple(_optional_int_list(_get_list(params, "vpn_forms"))),
        messenger_form_ids=tuple(_optional_int_list(_get_list(params, "messenger_forms"))),
        selection_mode=params.get("selection_mode") == "selected",
    )


def get_export_form_choices(user, filters: ExportFilters):
    vpn_forms = VpnTestForm.objects.select_related(
        "organization",
        "city",
        "mobile_operator",
        "created_by",
    ).filter(date__gte=filters.date_from, date__lte=filters.date_to)
    messenger_forms = MessengerTestForm.objects.select_related(
        "organization",
        "city",
        "mobile_operator",
        "created_by",
    ).filter(date__gte=filters.date_from, date__lte=filters.date_to)

    vpn_forms = _apply_common_form_filters(vpn_forms, filters, "")
    messenger_forms = _apply_common_form_filters(messenger_forms, filters, "")

    if filters.form_type == FORM_TYPE_VPN:
        messenger_forms = messenger_forms.none()
    elif filters.form_type == FORM_TYPE_MESSENGER:
        vpn_forms = vpn_forms.none()

    return {
        "vpn_forms": vpn_forms.order_by("date", "organization__name", "city__name", "mobile_operator__name", "os"),
        "messenger_forms": messenger_forms.order_by(
            "date", "organization__name", "city__name", "mobile_operator__name", "os"
        ),
    }


def build_excel_export(user, filters: ExportFilters) -> ExcelExportResult:
    workbook = Workbook()
    vpn_sheet = workbook.active
    vpn_sheet.title = "VPN"
    messenger_sheet = workbook.create_sheet("Мессенджеры")
    control_sheet = workbook.create_sheet("Контроль")

    rows_vpn = _write_vpn_sheet(vpn_sheet, filters) if filters.form_type in [FORM_TYPE_ALL, FORM_TYPE_VPN] else 0
    rows_messenger = (
        _write_messenger_sheet(messenger_sheet, filters)
        if filters.form_type in [FORM_TYPE_ALL, FORM_TYPE_MESSENGER]
        else 0
    )
    ready_for_export = _write_control_sheet(control_sheet, user, filters)

    for sheet in workbook.worksheets:
        _finalize_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return ExcelExportResult(
        content=buffer.getvalue(),
        file_name=_file_name(filters),
        rows_vpn=rows_vpn,
        rows_messenger=rows_messenger,
        ready_for_export=ready_for_export,
    )


def _write_vpn_sheet(sheet, filters: ExportFilters) -> int:
    headers = [
        "Дата",
        "Тестировщик",
        "Город",
        "ОС",
        "Сим-карта",
        "VPN",
        "Время замера",
        "Web без VPN",
        "Web через VPN",
        "Instagram",
        "YouTube",
        "Package ID / App ID",
        "Номер замера",
        "Комментарий",
        "Устройство",
    ]
    sheet.append(headers)

    qs = (
        VpnMeasurementResult.objects.select_related(
            "measurement",
            "measurement__form",
            "measurement__form__organization",
            "measurement__form__city",
            "measurement__form__mobile_operator",
            "measurement__form__created_by",
            "vpn_app",
        )
        .filter(measurement__form__date__gte=filters.date_from, measurement__form__date__lte=filters.date_to)
        .order_by(
            "measurement__form__date",
            "measurement__form__organization__name",
            "measurement__form__city__name",
            "measurement__form__mobile_operator__name",
            "measurement__form__os",
            "measurement__measurement_number",
            "id",
        )
    )
    qs = _apply_common_form_filters(qs, filters, "measurement__form")
    if filters.selection_mode:
        qs = qs.filter(measurement__form_id__in=filters.vpn_form_ids)

    count = 0
    for result in qs:
        form = result.measurement.form
        sheet.append(
            [
                _format_date(form.date),
                form.organization.name,
                form.city.name,
                form.os,
                form.mobile_operator.name,
                result.vpn_app.name,
                _format_datetime(result.measured_at),
                _binary_result(result.web_without_vpn_status),
                _binary_result(result.web_with_vpn_status),
                _binary_result(result.instagram_status),
                _binary_result(result.youtube_status),
                result.vpn_app.package_id,
                result.measurement.get_measurement_number_display(),
                result.comment,
                form.device,
            ]
        )
        count += 1
    return count


def _write_messenger_sheet(sheet, filters: ExportFilters) -> int:
    headers = [
        "Дата",
        "Тестировщик",
        "Город",
        "ОС",
        "Сим-карта",
        "Мессенджер",
        "Отправка сообщения",
        "Время отправки",
        "Объём файла",
        "Скорость отправки",
        "Аудиозвонок",
        "Комментарий",
        "Устройство",
    ]
    sheet.append(headers)

    qs = (
        MessengerTestResult.objects.select_related(
            "form",
            "form__organization",
            "form__city",
            "form__mobile_operator",
            "form__created_by",
            "messenger",
        )
        .filter(form__date__gte=filters.date_from, form__date__lte=filters.date_to)
        .order_by(
            "form__date",
            "form__organization__name",
            "form__city__name",
            "form__mobile_operator__name",
            "form__os",
            "id",
        )
    )
    qs = _apply_common_form_filters(qs, filters, "form")
    if filters.selection_mode:
        qs = qs.filter(form_id__in=filters.messenger_form_ids)

    count = 0
    for result in qs:
        form = result.form
        sheet.append(
            [
                _format_date(form.date),
                form.organization.name,
                form.city.name,
                form.os,
                form.mobile_operator.name,
                result.messenger.name,
                _binary_result(result.message_send_status),
                _format_time(result.message_sent_at),
                _decimal_or_blank(result.file_size_mb),
                _decimal_or_blank(result.file_send_speed_sec),
                _binary_result(result.audio_call_status),
                result.comment,
                form.device,
            ]
        )
        count += 1
    return count


def _write_control_sheet(sheet, user, filters: ExportFilters) -> bool:
    sheet.append(["Контроль"])
    if not filters.is_single_date:
        sheet.append(["Для периода контрольная сводка в MVP строится по начальной дате", _format_date(filters.date_from)])
    plans = plans_for_user(user, filters.date_from)
    if filters.organization_id:
        plans = plans.filter(organization_id=filters.organization_id)
    if filters.city_id:
        plans = plans.filter(city_id=filters.city_id)
    if filters.mobile_operator_id:
        plans = plans.filter(mobile_operator_id=filters.mobile_operator_id)
    if filters.os:
        plans = plans.filter(os=filters.os)
    if filters.form_type in [FORM_TYPE_VPN, FORM_TYPE_MESSENGER]:
        plans = plans.filter(form_type=filters.form_type)

    rows = build_plan_rows(plans)
    summary = split_summary(rows)
    groups = group_rows_by_organization(rows)

    _append_key_values(
        sheet,
        "Общая сводка",
        [
            ("Дата", _format_date(filters.date_from)),
            ("Всего ожидается", summary["total"]["expected"]),
            ("Всего создано", summary["total"]["created"]),
            ("Всего отсутствует", summary["total"]["missing"]),
            ("Черновиков", summary["total"]["draft"]),
            ("Частично отправлено", summary["total"]["partially_submitted"]),
            ("Отправлено", summary["total"]["submitted"]),
            ("Подтверждено", summary["total"]["confirmed"]),
            ("Возвращено", summary["total"]["returned"]),
            ("Готовность, %", summary["total"]["percent"]),
            ("Готов к выгрузке", "Да" if summary["ready_for_export"] else "Нет"),
        ],
    )
    _append_key_values(sheet, "VPN", _summary_values(summary["vpn"], include_partial=True))
    _append_key_values(sheet, "Мессенджеры", _summary_values(summary["messenger"], include_partial=False))

    sheet.append([])
    sheet.append(
        [
            "Организация",
            "Всего ожидается",
            "Отсутствует",
            "Черновик",
            "Частично отправлена",
            "Отправлена",
            "Подтверждена",
            "Возвращена",
            "Готовность, %",
        ]
    )
    for group in groups:
        group_summary = group["summary"]
        sheet.append(
            [
                group["organization"].name,
                group_summary["expected"],
                group_summary["missing"],
                group_summary["draft"],
                group_summary["partially_submitted"],
                group_summary["submitted"],
                group_summary["confirmed"],
                group_summary["returned"],
                group_summary["percent"],
            ]
        )
    return summary["ready_for_export"]


def _append_key_values(sheet, title, values):
    sheet.append([])
    sheet.append([title])
    sheet.append(["Показатель", "Значение"])
    for label, value in values:
        sheet.append([label, value])


def _summary_values(summary, include_partial):
    values = [
        ("Ожидается", summary["expected"]),
        ("Создано", summary["created"]),
        ("Отсутствует", summary["missing"]),
        ("Черновик", summary["draft"]),
    ]
    if include_partial:
        values.append(("Частично отправлена", summary["partially_submitted"]))
    values.extend(
        [
            ("Отправлена", summary["submitted"]),
            ("Подтверждена", summary["confirmed"]),
            ("Возвращена", summary["returned"]),
            ("Готовность, %", summary["percent"]),
        ]
    )
    return values


def _apply_common_form_filters(qs, filters: ExportFilters, prefix: str):
    field_prefix = f"{prefix}__" if prefix else ""
    if filters.organization_id:
        qs = qs.filter(**{f"{field_prefix}organization_id": filters.organization_id})
    if filters.city_id:
        qs = qs.filter(**{f"{field_prefix}city_id": filters.city_id})
    if filters.mobile_operator_id:
        qs = qs.filter(**{f"{field_prefix}mobile_operator_id": filters.mobile_operator_id})
    if filters.os:
        qs = qs.filter(**{f"{field_prefix}os": filters.os})
    if filters.status:
        qs = qs.filter(**{f"{field_prefix}status": filters.status})
    return qs


def _finalize_sheet(sheet):
    if sheet.max_row:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sheet.iter_rows():
        if row[0].value in ["Контроль", "Общая сводка", "VPN", "Мессенджеры"]:
            row[0].font = Font(bold=True)
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _file_name(filters: ExportFilters) -> str:
    if filters.is_single_date:
        return f"testing_export_{filters.date_from:%Y-%m-%d}.xlsx"
    return f"testing_export_{filters.date_from:%Y-%m-%d}_{filters.date_to:%Y-%m-%d}.xlsx"


def _parse_date(value):
    if not value:
        raise ValidationError("Укажите дату экспорта.")
    try:
        return date_class.fromisoformat(value)
    except ValueError as exc:
        raise ValidationError("Некорректная дата экспорта.") from exc


def _optional_int(value):
    if not value:
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError("Некорректный идентификатор фильтра.") from exc


def _optional_int_list(values):
    result = []
    for value in values:
        if value:
            result.append(_optional_int(value))
    return result


def _get_list(params, key):
    if hasattr(params, "getlist"):
        return params.getlist(key)
    value = params.get(key)
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def _label(value):
    return RESULT_LABELS.get(value, value or "")


def _binary_result(value):
    if value == CheckStatus.WORKS:
        return 1
    if value == CheckStatus.NOT_WORKS:
        return 0
    if value == CheckStatus.NOT_APPLICABLE:
        return "-"
    return _label(value)


def _form_status_label(value):
    return FORM_STATUS_LABELS.get(value, value or "")


def _format_date(value):
    return value.strftime("%d.%m.%Y") if value else ""


def _format_datetime(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime("%d.%m.%Y %H:%M")


def _format_time(value):
    return value.strftime("%H:%M") if value else ""


def _decimal_or_blank(value):
    return "" if value is None else float(value)

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from .models import OperatingSystem, VpnApp, VpnAppPeriod


@dataclass
class VpnImportResult:
    created: int = 0
    updated: int = 0
    deactivated: int = 0
    skipped: int = 0
    active_from: date | None = None
    active_to: date | None = None


def import_vpn_apps_from_excel(uploaded_file):
    from openpyxl import load_workbook

    workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    active_from, active_to = _extract_period(uploaded_file.name)
    if not active_from or not active_to:
        raise ValueError("в имени файла не найден период в формате 18.05-22.05")
    result = VpnImportResult(active_from=active_from, active_to=active_to)

    for sheet_name, os_name in _vpn_sheets(workbook):
        sheet = workbook[sheet_name]
        headers = _read_headers(sheet)
        imported_package_ids = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            payload = _row_to_payload(row, headers, os_name, active_from, active_to)
            if not payload:
                result.skipped += 1
                continue

            lookup = {"os": os_name, "package_id": payload.pop("package_id")}
            imported_package_ids.add(lookup["package_id"])
            period_payload = {
                "active_from": payload.pop("active_from"),
                "active_to": payload.pop("active_to"),
                "source_tag": payload.pop("source_tag"),
                "display_order": payload.pop("display_order"),
                "is_active": payload.pop("is_active"),
            }
            app, _ = VpnApp.objects.update_or_create(**lookup, defaults=payload)
            _, period_created = VpnAppPeriod.objects.update_or_create(
                vpn_app=app,
                active_from=period_payload["active_from"],
                active_to=period_payload["active_to"],
                defaults={
                    "source_tag": period_payload["source_tag"],
                    "display_order": period_payload["display_order"],
                    "is_active": True,
                },
            )
            if period_created:
                result.created += 1
            else:
                result.updated += 1
        if imported_package_ids:
            result.deactivated += VpnAppPeriod.objects.filter(
                vpn_app__os=os_name,
                active_from=active_from,
                active_to=active_to,
                is_active=True,
            ).exclude(vpn_app__package_id__in=imported_package_ids).update(is_active=False)

    return result


def export_vpn_periods_to_workbook(period_values):
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for os_name in (OperatingSystem.ANDROID, OperatingSystem.IOS):
        sheet = workbook.create_sheet(title=os_name)
        sheet.append(["Период", "Порядок", "Название VPN", "ID приложения", "Признак", "Ссылка на VPN", "Активно"])
        rows = (
            VpnAppPeriod.objects.select_related("vpn_app")
            .filter(vpn_app__os=os_name)
            .order_by("active_from", "display_order", "vpn_app__name")
        )
        if period_values:
            rows = rows.filter(_period_query(period_values))
        for item in rows:
            sheet.append(
                [
                    _format_period(item.active_from, item.active_to),
                    item.display_order,
                    item.vpn_app.name,
                    item.vpn_app.package_id,
                    item.source_tag,
                    item.vpn_app.store_url,
                    "Да" if item.is_active else "Нет",
                ]
            )
        for column in ("A", "C", "D", "E"):
            sheet.column_dimensions[column].width = 32

    return workbook


def _period_query(period_values):
    from django.db.models import Q

    query = Q()
    for value in period_values:
        active_from, active_to = _parse_period_value(value)
        query |= Q(active_from=active_from, active_to=active_to)
    return query


def _parse_period_value(value):
    active_from, active_to = value.split("__", 1)
    return date.fromisoformat(active_from), date.fromisoformat(active_to)


def _format_period(active_from, active_to):
    return f"{active_from:%d.%m.%Y}-{active_to:%d.%m.%Y}"


def _vpn_sheets(workbook):
    for sheet_name in workbook.sheetnames:
        normalized = sheet_name.lower().replace(" ", "")
        if "android" in normalized:
            yield sheet_name, OperatingSystem.ANDROID
        elif "ios" in normalized:
            yield sheet_name, OperatingSystem.IOS


def _read_headers(sheet):
    headers = {}
    for index, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))):
        if value:
            headers[_normalize_header(value)] = index
    return headers


def _row_to_payload(row, headers, os_name, active_from, active_to):
    order = _get(row, headers, "номер", "№")
    name = _get(row, headers, "название vpn")
    store_url = _get(row, headers, "ссылка на vpn")
    source_tag = _get(row, headers, "признак")
    package_id = _get(row, headers, "id приложения", "id")

    if not name or not package_id:
        return None

    return {
        "name": str(name).strip(),
        "store_url": str(store_url).strip() if store_url else "",
        "package_id": str(package_id).strip(),
        "source_tag": str(source_tag).strip() if source_tag else "",
        "is_active": True,
        "active_from": active_from,
        "active_to": active_to,
        "display_order": _to_int(order, default=100),
    }


def _get(row, headers, *names):
    for name in names:
        index = headers.get(_normalize_header(name))
        if index is not None and index < len(row):
            return row[index]
    return None


def _normalize_header(value):
    return str(value).strip().lower()


def _to_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _extract_period(filename):
    match = re.search(r"(\d{2})\.(\d{2})\s*-\s*(\d{2})\.(\d{2})", Path(filename).name)
    if not match:
        return None, None

    start_day, start_month, end_day, end_month = map(int, match.groups())
    year = date.today().year
    return date(year, start_month, start_day), date(year, end_month, end_day)
